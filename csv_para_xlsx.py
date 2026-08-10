r"""Converte consultas.csv (UTF-8, virgula, campos multilinha) para .xlsx em modo tabela.

Uso:
    .\.venv\Scripts\python.exe csv_para_xlsx.py [entrada.csv] [saida.xlsx]
"""

import csv
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

LARGURA_MAX = 55


def converter(entrada: Path, saida: Path) -> int:
    # utf-8-sig descarta o BOM caso exista; newline="" deixa o csv tratar
    # as quebras de linha dentro dos campos entre aspas.
    with entrada.open(encoding="utf-8-sig", newline="") as f:
        linhas = list(csv.reader(f))

    if not linhas:
        raise SystemExit(f"{entrada} esta vazio")

    wb = Workbook()
    ws = wb.active
    ws.title = "consultas"

    for linha in linhas:
        ws.append(linha)

    n_linhas = len(linhas)
    n_colunas = len(linhas[0])
    ref = f"A1:{get_column_letter(n_colunas)}{n_linhas}"

    # Tabela nativa do Excel: filtro, faixas alternadas e cabecalho fixo.
    tabela = Table(displayName="Consultas", ref=ref)
    tabela.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9", showRowStripes=True
    )
    ws.add_table(tabela)
    ws.freeze_panes = "A2"

    cabecalho = PatternFill("solid", fgColor="DDEBF7")
    for celula in ws[1]:
        celula.font = Font(bold=True)
        celula.fill = cabecalho

    for i in range(1, n_colunas + 1):
        letra = get_column_letter(i)
        maior = max(
            (len(str(c.value)) for c in ws[letra] if c.value is not None),
            default=10,
        )
        ws.column_dimensions[letra].width = min(maior + 2, LARGURA_MAX)

    # So as colunas largas quebram linha; o resto fica compacto.
    for linha in ws.iter_rows(min_row=2):
        for celula in linha:
            if celula.column_letter and ws.column_dimensions[
                celula.column_letter
            ].width >= LARGURA_MAX:
                celula.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                celula.alignment = Alignment(vertical="top")

    wb.save(saida)
    return n_linhas - 1


if __name__ == "__main__":
    base = Path(__file__).parent
    entrada = Path(sys.argv[1]) if len(sys.argv) > 1 else base / "consultas.csv"
    saida = Path(sys.argv[2]) if len(sys.argv) > 2 else entrada.with_suffix(".xlsx")

    total = converter(entrada, saida)
    print(f"{total} registros -> {saida}")
